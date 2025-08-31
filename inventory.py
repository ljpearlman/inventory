from openpyxl import Workbook, load_workbook
import sys
import csv
import requests
from html.parser import HTMLParser
from pathlib import Path
import urllib.parse

class ContentsRow:
    attributes = [
        'name',
        'quantity',
        'servpro_col',
        'servpro_icat_id',
        'image_url',
        'laura_cost_estimate',
        'purchase_date',
        'aaa_estimate',
        'orig_receipt',
        'replacement_receipt',
        'reimbursed_to_date',
        'comments'
    ]

       
    def __init__(self, values):
        self.values = values

class ServproSheet:
    headers = [
        'servpro_col',
        'servpro_icat_id',        
        'name',
        'quantity',
        'image_url'
    ]
    
    def __init__(self, filename):
        self.excel_rows = []
        self.rows = []
        wb = load_workbook(filename=filename)
        sheet = wb['Schedule of Loss']
        for excel_row in sheet.iter_rows():
            if len(excel_row) > 0 and isinstance(excel_row[0].value, int):
                self.excel_rows.append(excel_row)

    def parse(self):
        for excel_row in self.excel_rows:
            row = {
                'servpro_col' : excel_row[0].value,
                'servpro_icat_id' : excel_row[1].value,
                'image_url' : excel_row[2].hyperlink.target,
                'quantity' : excel_row[3].value.split(' ', 1)[0],
                'name' : excel_row[3].value.split(' ', 1)[1]
                }
            self.rows.append(row)

    def populate_images(self, row, cachedir):
        parser = MyHTMLParser(row, cachedir)
        r = requests.get(row['image_url'])        
        parser.feed(r.text)

    def print(self):
        writer = csv.DictWriter(sys.stdout, self.headers)
        writer.writeheader()
        for row in self.rows:
            writer.writerow(row)

class MyHTMLParser(HTMLParser):
    def __init__(self, row, cachedir):
        self.basename = f"sp-{row['servpro_icat_id']}"
        self.urlbase = row['image_url']
        self.dir = Path(cachedir) / self.basename
        self.dir.mkdir(exist_ok=True)
        super(MyHTMLParser, self).__init__()
        self.stage = None
        self.ordinal = 0
        
    def handle_starttag(self, tag, attrs):
        if tag == 'a':
            self.in_image = False
            for attr in attrs:
                if attr[0] == 'data-fancybox':
                    self.in_image = True
                    self.ordinal = self.ordinal + 1
                    break

            if self.in_image:
                for attr in attrs:
                    if attr[0] == 'href':
                        file = self.dir / f'{self.basename}-{self.ordinal:02}'
                        url = urllib.parse.urljoin(self.urlbase, attr[1])
                        r = requests.get(url)
                        file.write_bytes(r.content)
                        

    def handle_endtag(self, tag):
        if tag == 'a':
            self.in_image = False

    def handle_data(self, data):
        pass

if __name__ == '__main__':
    sheet = ServproSheet('original.xlsm')
    sheet.parse()
    for row in sheet.rows:
        sheet.populate_images(row)
#    sheet.print()
